# ==============================================================
# 6단계: 네이버 뉴스 스크래핑 (실전) 📰
# ==============================================================
# 
# 🎯 목표:
#    - 네이버 뉴스(IT/과학) 섹션의 헤드라인 뉴스 수집
#    - 제목, 내용 요약, 신문사, 링크 정보 가져오기
#    - 엑셀 파일로 저장하고 예쁘게 꾸미기 (스타일링)
#
# ==============================================================

import requests
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --------------------------------------------------------------
# 🌐 1. 웹사이트 접속 (User-Agent 설정 필수!)
# --------------------------------------------------------------
url = "https://news.naver.com/section/105"  # IT/과학 뉴스

# 🚨 중요: 네이버 같은 대형 포털은 로봇(프로그램) 접속을 막을 수 있습니다.
# "나는 로봇이 아니라 브라우저야!"라고 알려주기 위해 User-Agent 정보를 함께 보냅니다.
headers = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
}

print(f"🌐 네이버 뉴스에 접속 중... ({url})")
response = requests.get(url, headers=headers)

# 접속 확인
if response.status_code != 200:
    print(f"❌ 접속 실패! 상태 코드: {response.status_code}")
    exit()

# --------------------------------------------------------------
# 🥣 2. HTML 분석 준비
# --------------------------------------------------------------
soup = BeautifulSoup(response.text, "html.parser")

# --------------------------------------------------------------
# 🔍 3. 뉴스 데이터 찾기
# --------------------------------------------------------------
print("🔍 뉴스를 찾고 있습니다...")

# 데이터를 담을 리스트
news_list = []

# 네이버 뉴스 리스트 아이템 찾기
news_items = soup.select("li.sa_item")

print(f"총 {len(news_items)}개의 뉴스를 찾았습니다.")

for item in news_items:
    try:
        # news_list에 담을 딕셔너리 생성
        news_data = {}

        # 1) 뉴스 제목
        title_tag = item.select_one("strong.sa_text_strong")
        if not title_tag: continue
        news_data["뉴스 제목"] = title_tag.text.strip()

        # 2) 뉴스 내용 요약
        summary_tag = item.select_one("div.sa_text_lede")
        news_data["뉴스 내용"] = summary_tag.text.strip() if summary_tag else "요약 없음"

        # 3) 신문사
        press_tag = item.select_one("div.sa_text_press")
        news_data["신문사"] = press_tag.text.strip() if press_tag else "알 수 없음"

        # 4) 뉴스 링크
        link_tag = item.select_one("a.sa_text_title")
        news_data["링크"] = link_tag["href"] if link_tag else ""
        
        news_list.append(news_data)

    except Exception as e:
        print(f"⚠️ 에러 발생: {e}")
        continue

print(f"✅ 수집 완료: {len(news_list)}개")

# --------------------------------------------------------------
# 💾 4. 엑셀 파일로 저장하기
# --------------------------------------------------------------
today = datetime.now()
date_str = today.strftime("%Y_%m_%d")
file_name = f"naver_news_{date_str}.xlsx"

print(f"💾 엑셀 파일로 저장 중... ({file_name})")

df = pd.DataFrame(news_list)
df.to_excel(file_name, index=False)

# --------------------------------------------------------------
# 🎨 5. 엑셀 파일 꾸미기 (스타일링)
# --------------------------------------------------------------
print("🎨 엑셀 파일을 예쁘게 꾸미는 중...")

# 저장된 엑셀 파일 불러오기
wb = load_workbook(file_name)
ws = wb.active  # 현재 활성화된 시트 선택

# 1) 열 너비 조절
ws.column_dimensions['A'].width = 50  # 뉴스 제목
ws.column_dimensions['B'].width = 80  # 뉴스 내용
ws.column_dimensions['C'].width = 15  # 신문사
ws.column_dimensions['D'].width = 40  # 링크

# 2) 스타일 정의
# 헤더 스타일
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")

# 본문 스타일
body_font = Font(size=11)
body_align = Alignment(vertical="center", wrap_text=True) # wrap_text=True: 줄바꿈 허용
link_font = Font(color="0000FF", underline="single") # 링크 스타일

# 테두리 스타일
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# 3) 헤더(첫 번째 줄) 스타일 적용
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# 4) 본문(데이터) 스타일 적용
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = body_font
        cell.alignment = body_align
        cell.border = thin_border
        
        # 링크 컬럼(4번째(D) 열)에 하이퍼링크 스타일 적용
        if cell.column == 4 and cell.value:
            cell.font = link_font

# 5) 변경사항 저장
wb.save(file_name)

print("🎉 모든 작업 완료!")
